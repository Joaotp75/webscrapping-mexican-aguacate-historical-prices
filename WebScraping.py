from bs4 import BeautifulSoup
import requests
import pandas as pd
from openpyxl import Workbook
from datetime import datetime
from io import StringIO
from openpyxl.utils.dataframe import dataframe_to_rows
from icecream import ic

# Criando um dicionario com os tipos de Aguacate Hass e seus respectivos codigos na url das consultas.
Produtos = {'Sin Classificar': 132,
            'Primeira': 133,
            'Segunda': 134,
            'Adelantado': 135,
            'Calidad Extra': 136,
            'Calidad Super Extra': 137,
            'Flor Vieja': 138}

# Essa função faz uma requisição HTTP GET para uma URL específica, que é gerada pela função _get_url, usando os códigos dos produtos. Ela coleta os dados de preços para um produto num intervalo de datas definido.
def get_data(codigo_produto: int, year: str = '2000'):
    data_inicio = f'{year}-01-01'
    data_fim = f'{year}-12-31'
    return requests.get(_get_url(codigo_produto, data_inicio=data_inicio, data_fim=data_fim))

# Constrói a URL para a requisição com base no código do produto e nas datas de início e fim.
def _get_url(codigo_produto: int, data_inicio: str = '2000-01-01',
             data_fim: str | datetime = datetime.today()):
    data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')
    data_fim = datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')
    return f'http://www.economia-sniim.gob.mx/nuevo/Consultas/MercadosNacionales/' \
           f'PreciosDeMercado/Agricolas/ResultadosConsultaFechaFrutasYHortalizas.aspx?' \
           f'fechaInicio={data_inicio}&fechaFinal={data_fim}&ProductoId={codigo_produto}&' \
           f'OrigenId=-1&Origen=Todos&DestinoId=-1&Destino=Todos&PreciosPorId=1' \
           f'&RegistrosPorPagina=100000'

#  Extraindo o peso para padronizar os preços por kg.
def _get_kg(presentacion: str):
  if presentacion == 'Kilogramo':
    return 1
  else:
    try:
      return int(presentacion.split(' ')[2])
    except:
      return 0 # a medida 'manojo', que significa punhado, é dificl de saber a kg correta


 # Converte um DataFrame do pandas em linhas e colunas de uma planilha Excel.
def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    rows = dataframe_to_rows(df, header=header, index=index)
    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
             ws.cell(row=r_idx, column=c_idx).value = value

wb = Workbook()
resumo = wb.create_sheet('Médias mensais', 0)
col = ['A', 'F', 'K', 'P', 'U','Z','AE']
number = 0


# Converter a letra identicadora de colunas no excel para o número correspondente índice númerico.
def column_index_from_string(col_str):
    number = 0
    for char in col_str:
        number = number * 26 + (ord(char.upper()) - ord('A')) + 1
    return number

# Loop principal, que itera as consultas dos tipos de Aguacate Hass
for produto, codigo in Produtos.items():
    ic(produto)
    years = range(1997, 2025)
    dfs = []
    #Loop secundário que itera os intervalos de 1 ano na janela desejada para fazer "microconsultas", já que a grande para o tipo de produto "primeira" estava ficando grande demais e não funcionando. E fica prevenido o caso de não ter tabela com id='tblResultados' que acontece quando não há dados para aquele produto naquele intervalo de 1 ano.
    for year in years:
      ic(year)
      soup = BeautifulSoup(get_data(codigo, str(year)).text, 'html.parser')
      table = soup.find_all('table', id='tblResultados')
      try:
        df = pd.read_html(StringIO(str(table)), header=0)[0]
      except:
        df = pd.DataFrame()
      dfs += df,
    df = pd.concat(dfs, axis=0)
    df = df.drop(0, axis=0)
    # Organizando as tabelas para ter valores númericos no preços, ordenar as linhas por data e padronizar os preços "/Kg"
    df['Fecha'] = pd.to_datetime(df.Fecha, format='%d/%m/%Y')
    df['Precio Mín'] = pd.to_numeric(df['Precio Mín'])
    df['Precio Max'] = pd.to_numeric(df['Precio Max'])
    df['Precio Frec'] = pd.to_numeric(df['Precio Frec'])
    df['Kg'] = df['Presentación'].apply(_get_kg)
    df['Precio Mín/Kg'] = df['Precio Mín'] / df['Kg']
    df['Precio Max/Kg'] = df['Precio Max'] / df['Kg']
    df['Precio Frec/Kg'] = df['Precio Frec'] / df['Kg']
    df.drop(columns=['Origen', 'Destino', 'Obs.', 'Presentación'], inplace=True)
    df = df.sort_values(by='Fecha')

    #Criar a planilha dos dados brutos de cada produto
    ws1 = wb.create_sheet(f"{produto}_bruto")
    df_to_excel(df, ws1)
    # Criando a aba de médias mensais na planilha
    d2 = df.copy()
    d2['Fecha'] = d2['Fecha'].dt.strftime('%Y/%m')
    d2 = d2.groupby(['Fecha'])[['Precio Mín/Kg', 'Precio Max/Kg', 'Precio Frec/Kg']].mean()

    col_number = column_index_from_string(col[number])

    resumo.cell(row=1, column=col_number).value = produto

    df_to_excel(d2, resumo, startcol=col_number, startrow=1)

    number += 1

# Removendo a planilha vazia e salvando o arquivo
sheet = wb['Sheet']
wb.remove(sheet)
wb.save("resultado.xlsx")
